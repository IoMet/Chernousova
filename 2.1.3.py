import csv
import re
from datetime import datetime
import os
from statistics import mean
import openpyxl
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
import pdfkit
from jinja2 import Environment, FileSystemLoader

currency_to_rub = {
    "Манаты": 35.68,
    "Белорусские рубли": 23.91,
    "Евро": 59.90,
    "Грузинский лари": 21.74,
    "Киргизский сом": 0.76,
    "Тенге": 0.13,
    "Рубли": 1,
    "Гривны": 1.64,
    "Доллары": 60.66,
    "Узбекский сум": 0.0055,
}

dic_currency = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум",
}


def make_exit(msg):
    print(msg)
    exit(0)


class UserInput:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')


class DataSet:
    def __init__(self, file_name):
        if os.stat(file_name).st_size == 0:
            make_exit("Пустой файл")

        self.vacancies = [row for row in csv.reader(open(file_name, encoding="utf_8_sig"))]
        self.titles = self.vacancies[0]
        self.parsed_vacancies = [row for row in self.vacancies[1:] if
                                 len(row) == len(self.titles) and row.count('') == 0]

        if len(self.parsed_vacancies) == 0:
            make_exit('Нет данных')


class Vacancy:
    def __init__(self, one_vac):
        for key, value in one_vac.items():
            if key == 'salary_currency':
                self.salary_currency = dic_currency[value]
            elif key == "salary_to":
                self.salary_to = '{:,}'.format(int(float(value))).replace(',', ' ')
            elif key == "salary_from":
                self.salary_from = '{:,}'.format(int(float(value))).replace(',', ' ')
            elif key == 'published_at':
                self.published_at = int(datetime.strptime(value, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
            elif key == "name":
                self.name = value
            elif key == "area_name":
                self.area_name = value

        self.salary = int(currency_to_rub[self.salary_currency] * (
                float(self.salary_from.replace(' ', '')) + float(self.salary_to.replace(' ', ''))) // 2)

    def fill_data_dicts(self, salary_dicts, count_dicts, vacancy):
        year = self.published_at
        area = self.area_name

        count_dicts.vac_count[year] += 1
        salary_dicts.salaries[year] += [self.salary]

        if area in count_dicts.vacancies_areas.keys():
            count_dicts.vacancies_areas[area] += 1
        else:
            count_dicts.vacancies_areas[area] = 1

        if area in salary_dicts.salaries_areas.keys():
            salary_dicts.salaries_areas[area] += [self.salary]
        else:
            salary_dicts.salaries_areas[area] = [self.salary]

        if vacancy in self.name and vacancy != "":
            count_dicts.vacancy_count[year] += 1
            salary_dicts.vacancy_salaries[year] += [self.salary]

        salary_dicts.vacancies += 1


class StaticSalary:
    salaries: dict = {}
    vacancy_salaries: dict = {}
    salaries_areas: dict = {}
    vacancies: int = 0

    def __init__(self):
        for i in range(2007, 2023):
            self.salaries[i] = []
            self.vacancy_salaries[i] = []

    @staticmethod
    def check_len_dic(dic, vacancy=""):
        new_dic = {}
        for name, item in dic.items():
            if len(item) != 0:
                new_dic[name] = int(mean(item))
        if vacancy != "" and len(new_dic) == 0:
            return {2022: 0}
        return new_dic


class StaticCount:
    vac_count: dict = {}
    vacancy_count: dict = {}
    vacancies_areas: dict = {}

    def __init__(self):
        for i in range(2007, 2023):
            self.vac_count[i] = 0
            self.vacancy_count[i] = 0

    @staticmethod
    def check_int_dic(dic, vacancy=""):
        new_dic = {}
        for name, item in dic.items():
            if item != 0:
                new_dic[name] = item
        if vacancy != "" and len(new_dic) == 0:
            return {2022: 0}
        return new_dic


class ResultStatic:
    list_by_year: list
    list_by_area: list
    others: int

    def check_one_percent(self, salaries, counts):
        vacancies_areas_dict = {}
        salaries_areas_dict = {}
        self.others = 0
        for key, value in counts.vacancies_areas.items():
            percent = value / salaries.vacancies
            if percent >= 0.01:
                salaries_areas_dict[key] = int(mean(salaries.salaries_areas[key]))
                vacancies_areas_dict[key] = round(percent, 4)
            else:
                self.others += percent
        salaries.salaries_areas = salaries_areas_dict
        counts.vacancies_areas = vacancies_areas_dict

    @staticmethod
    def sort_dict(dictionary):
        sorted_tuples = sorted(dictionary.items(), key=lambda item: item[1], reverse=True)[:10]
        sorted_dict = {k: v for k, v in sorted_tuples}
        return sorted_dict

    def print_result(self, vacancy, salaries, counts):
        self.check_one_percent(salaries, counts)
        format_salaries = salaries.check_len_dic(salaries.salaries)
        format_vac_count = counts.check_int_dic(counts.vac_count)
        format_vacancy_salaries = salaries.check_len_dic(salaries.vacancy_salaries, vacancy)
        format_vacancy_count = counts.check_int_dic(counts.vacancy_count, vacancy)
        format_salaries_areas = self.sort_dict(salaries.salaries_areas)
        format_vacancies_areas = self.sort_dict(counts.vacancies_areas)

        print("Динамика уровня зарплат по годам:", format_salaries)
        print("Динамика количества вакансий по годам:", format_vac_count)
        print("Динамика уровня зарплат по годам для выбранной профессии:", format_vacancy_salaries)
        print("Динамика количества вакансий по годам для выбранной профессии:", format_vacancy_count)
        print("Уровень зарплат по городам (в порядке убывания):", format_salaries_areas)
        print("Доля вакансий по городам (в порядке убывания):", format_vacancies_areas)

        self.list_by_year = [format_salaries, format_vacancy_salaries,
                             format_vac_count, format_vacancy_count]
        self.list_by_area = [format_salaries_areas, format_vacancies_areas]


class Report:
    rows_by_year = ["Год", "Средняя зарплата", "Средняя зарплата - ", "Количество вакансий",
                    "Количество вакансий - "]
    rows_by_area = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]

    def __init__(self, vacancy_name, list_by_year, list_by_area, others):
        self.book = openpyxl.Workbook()
        self.book.remove(self.book.active)

        self.page_by_year = self.book.create_sheet("Статистика по годам")
        self.page_by_area = self.book.create_sheet("Статистика по городам")

        self.generate_excel(vacancy_name, list_by_year, list_by_area)
        self.generate_image(vacancy_name, list_by_year, list_by_area, others)
        self.generate_pdf(vacancy_name, list_by_year, list_by_area)

    @staticmethod
    def generate_pdf(vacancy_name, list_by_year, list_by_area):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("2.1.3.html")

        pdf_template = template.render(
            {'name': vacancy_name, 'list_by_year': list_by_year, 'list_by_area': list_by_area,
             'area_td_1': list(list_by_area[0].keys()), 'area_td_2': list(list_by_area[0].values()),
             'area_td_3': list(list_by_area[1].keys()), 'area_td_4': list(list_by_area[1].values())})

        options = {'enable-local-file-access': None}
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    @staticmethod
    def generate_image(vacancy_name, list_by_year, list_by_area, others):

        x_nums = np.arange(len(list_by_year[0].keys()))
        width = 0.4
        x_place1 = x_nums - width / 2
        x_place2 = x_nums + width / 2
        fig = plt.figure()

        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_place1, list_by_year[0].values(), width, label="средняя з/п")
        ax.bar(x_place2, list_by_year[1].values(), width, label=f"з/п {vacancy_name}")
        ax.legend(fontsize=8)
        ax.set_xticks(x_nums, list_by_year[0].keys(), rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам")
        ax.bar(x_place1, list_by_year[2].values(), width, label="Количество вакансий")
        ax.bar(x_place2, list_by_year[3].values(), width, label=f"Количество вакансий \n{vacancy_name}")
        ax.tick_params(axis="both", labelsize=8)
        ax.set_xticks(x_nums, list_by_year[2].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        width = 0.8
        y_ticks_cities = np.arange(len(list_by_area[0].keys()))
        y_ticks_cities_names = {}
        for key, value in list_by_area[0].items():
            if "-" in key or " " in key:
                key = key.replace("-", "-\n")
                key = key.replace(" ", "\n")
            y_ticks_cities_names[key] = value

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        ax.barh(y_ticks_cities, list_by_area[0].values(), width, align="center")
        ax.xaxis.set_major_locator(ticker.MultipleLocator(40000))
        ax.set_yticks(y_ticks_cities, labels=y_ticks_cities_names.keys(), horizontalalignment="right",
                      verticalalignment="center")
        ax.tick_params(axis="x", labelsize=8)
        ax.tick_params(axis="y", labelsize=6)
        ax.invert_yaxis()
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        list_by_area[1]["Другие"] = others
        ax.pie(list_by_area[1].values(), labels=list_by_area[1].keys(), textprops={'size': 6})
        ax.axis('equal')

        plt.tight_layout()
        plt.savefig("graph.png")

    def generate_excel(self, vacancy_name, list_by_year, list_by_area):
        thins = Side(border_style="thin", color="000000")

        self.set_value_page(vacancy_name, list_by_year, list_by_area, thins)
        self.set_width(thins)

        self.book.save("report.xlsx")

    def set_value_page(self, vacancy_name, list_by_year, list_by_area, thins):
        # Для первого листа
        for i, value in enumerate(self.rows_by_year, 1):
            self.page_by_year.cell(row=1, column=i).value = value + vacancy_name if " - " in value else value

        for year, value in list_by_year[0].items():
            self.page_by_year.append([year, value, list_by_year[1][year], list_by_year[2][year], list_by_year[3][year]])

        # Для второго листа
        for i, value in enumerate(self.rows_by_area, 1):
            self.page_by_area.cell(row=1, column=i).value = value

        for i in range(len(list_by_area[0])):
            self.page_by_area.append([list(list_by_area[0].keys())[i],
                                      list(list_by_area[0].values())[i], "",
                                      list(list_by_area[1].keys())[i],
                                      list(list_by_area[1].values())[i]])

    def set_width(self, thins):
        dims = {}
        for row in self.page_by_year.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.page_by_year.column_dimensions[col].width = value + 2

        dims = {}
        for row in self.page_by_area.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.page_by_area.column_dimensions[col].width = value + 2

        # Изменение ячеек страниц
        for i in range(17):
            for j in range(5):
                self.page_by_year.cell(row=i + 1, column=j + 1).border = Border(top=thins, bottom=thins, left=thins,
                                                                                right=thins)
        for i in range(11):
            for j in range(5):
                if j != 2:
                    self.page_by_area.cell(row=i + 1, column=j + 1).border = Border(top=thins, bottom=thins,
                                                                                    left=thins,
                                                                                    right=thins)
        for i in range(10):
            self.page_by_area.cell(row=i + 2, column=5).number_format = "0.00%"

        for i in range(5):
            self.page_by_year.cell(row=1, column=i + 1).font = Font(bold=True)
            self.page_by_area.cell(row=1, column=i + 1).font = Font(bold=True)


class PrepareVacData:
    def __init__(self):
        salaries = StaticSalary()
        counts = StaticCount()
        result = ResultStatic()

        for one_parsed_vac in parsed_vac:
            parsed_data = Vacancy(dict(zip(titles, map(self.prepare, one_parsed_vac))))
            parsed_data.fill_data_dicts(salaries, counts, inputed.vacancy_name)
        result.print_result(inputed.vacancy_name, salaries, counts)
        report = Report(inputed.vacancy_name, result.list_by_year, result.list_by_area, result.others)

    @staticmethod
    def prepare(text):
        text = re.sub('<.*?>', '', text)
        text = text.replace("\r\n", "\n")
        res = [' '.join(word.split()) for word in text.split('\n')]
        return res[0] if len(res) == 1 else res


inputed = UserInput()
dataset = DataSet(inputed.file_name)

titles = dataset.titles
parsed_vac = dataset.parsed_vacancies

prepare_vac = PrepareVacData()
